# =============================================================================
# AutoInsight AI — Report Phase 4: Assembly & Multi-Format Export (Phase 3 Full)
# Phase 3: Report Engine — HTML/Markdown/PDF/Excel Export
# =============================================================================
"""
Phase 4: Assembly & Multi-Format Export — Fully Deterministic.

Assembles validated report sections into a complete ReportBundle and
exports to ALL 4 supported formats: HTML, Markdown, PDF, Excel.

This phase is FULLY DETERMINISTIC — no LLM calls.

Export Engines:
  1. HTML:    Jinja2 template → professional HTML page (with CSS styling)
  2. Markdown: Jinja2 template → clean .md file
  3. PDF:     weasyprint HTML→PDF conversion (or markdown→pdf fallback)
  4. Excel:   OpenPyXL workbook with structured sheets

Storage:
  - Exports stored in S3/MinIO with signed URLs
  - Metadata indexed in PostgreSQL `reports` table
  - Returns {report_id, export_urls: {pdf, html, md, xlsx}}

Usage:
    from backend.report.phase4_export import Phase4_Export
    exporter = Phase4_Export()
    bundle = await exporter.run(sections)
    urls = await exporter.export_all(bundle)
"""

from __future__ import annotations

import io
import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

from jinja2 import Environment, FileSystemLoader
from pydantic import ValidationError

from backend.cache import cache_manager
from backend.config import settings
from backend.database import insert_one
from backend.schemas import (
    ReportBundle,
    ReportSection,
    ConfidenceLevel,
)
from backend.storage import storage_manager

logger = logging.getLogger(__name__)

# Template directory
TEMPLATE_DIR = Path(__file__).parent / "templates"


class Phase4_Export:
    """
    Phase 4: Assembly & Multi-Format Export.
    
    Assembles validated report sections into a ReportBundle and exports
    to all 4 supported formats. Fully deterministic — no LLM calls.
    
    Export formats:
      - HTML: Styled professional document with confidence badges
      - Markdown: Clean GitHub-flavored markdown
      - PDF: Print-ready document (via weasyprint HTML→PDF)
      - Excel: Structured workbook with multiple sheets
    """

    def __init__(self):
        """Initialize the export engine with Jinja2 templates."""
        self.env = Environment(
            loader=FileSystemLoader(str(TEMPLATE_DIR)),
            autoescape=True,
        )

    # ─── Phase 4 Main Entry Point ─────────────────────────────────────────

    async def run(
        self,
        sections: List[ReportSection],
        title: Optional[str] = None,
        audit_trail: Optional[List[Dict[str, Any]]] = None,
        viz_payload: Optional[Dict[str, Any]] = None,
    ) -> ReportBundle:
        """
        Assemble all sections into a complete ReportBundle.
        
        Args:
            sections: Validated sections from Phase 3
            title: Optional report title (auto-generated if None)
            audit_trail: Optional audit trail entries
            viz_payload: Optional visualization configuration
        
        Returns:
            Complete ReportBundle ready for export
        """
        logger.info(f"Report Phase 4: Assembling {len(sections)} sections into ReportBundle")

        bundle = ReportBundle(
            title=title or "Auto-Generated Analytical Report",
            sections=sections,
            overall_confidence=self._compute_overall_confidence(sections),
            audit_trail=audit_trail or [],
            export_metadata={
                "formats": ["pdf", "html", "md", "xlsx"],
                "generated_at": datetime.utcnow().isoformat(),
                "section_count": len(sections),
                "export_status": {
                    "pdf": "pending",
                    "html": "pending",
                    "md": "pending",
                    "xlsx": "pending",
                },
            },
            viz_payload=viz_payload or self._build_viz_payload(sections),
        )

        logger.info(
            f"Report Phase 4: Assembly complete — "
            f"report_id={bundle.report_id}, "
            f"overall_confidence={bundle.overall_confidence:.2f}"
        )

        return bundle

    # ─── Export All Formats ───────────────────────────────────────────────

    async def export_all(
        self,
        bundle: ReportBundle,
    ) -> Dict[str, str]:
        """
        Export report to all 4 formats and store in S3/MinIO.
        
        Args:
            bundle: Complete ReportBundle to export
        
        Returns:
            Dict mapping format names to storage URLs
        """
        logger.info(f"Exporting report {bundle.report_id[:12]}... to all formats")

        export_urls = {}
        base_path = f"reports/{bundle.report_id}"

        # Export each format
        try:
            # HTML export
            html_content = await self.export_html(bundle)
            html_key = f"{base_path}/report.html"
            await self._store_export(html_key, html_content.encode(), "text/html")
            export_urls["html"] = html_key

            # Markdown export
            md_content = await self.export_markdown(bundle)
            md_key = f"{base_path}/report.md"
            await self._store_export(md_key, md_content.encode(), "text/markdown")
            export_urls["md"] = md_key

            # PDF export (from HTML)
            try:
                pdf_key = await self.export_pdf(bundle)
                export_urls["pdf"] = pdf_key
            except Exception as e:
                logger.warning(f"PDF export failed: {e}")
                export_urls["pdf"] = None

            # Excel export
            try:
                xlsx_key = await self.export_excel(bundle)
                export_urls["xlsx"] = xlsx_key
            except Exception as e:
                logger.warning(f"Excel export failed: {e}")
                export_urls["xlsx"] = None

            # Persist to PostgreSQL
            await self._persist_report(bundle, export_urls)

            # Update cache
            await cache_manager.set(
                f"report:{bundle.report_id}",
                {
                    "report_id": bundle.report_id,
                    "title": bundle.title,
                    "overall_confidence": bundle.overall_confidence,
                    "export_urls": export_urls,
                    "generated_at": bundle.generated_at.isoformat(),
                },
                ttl=86400,
            )

            logger.info(
                f"Report {bundle.report_id[:12]}... exported: "
                f"html={'✅' if export_urls.get('html') else '❌'} "
                f"md={'✅' if export_urls.get('md') else '❌'} "
                f"pdf={'✅' if export_urls.get('pdf') else '❌'} "
                f"xlsx={'✅' if export_urls.get('xlsx') else '❌'}"
            )

        except Exception as e:
            logger.error(f"Export failed: {e}")

        return export_urls

    # ─── 1. HTML Export (Jinja2) ──────────────────────────────────────────

    async def export_html(self, bundle: ReportBundle) -> str:
        """
        Export report as a styled HTML document using Jinja2 template.
        
        Returns:
            Rendered HTML string
        """
        template = self.env.get_template("report_html.jinja2")
        html = template.render(report=bundle)
        logger.info(f"HTML export: {len(html)} chars")
        return html

    # ─── 2. Markdown Export (Jinja2) ──────────────────────────────────────

    async def export_markdown(self, bundle: ReportBundle) -> str:
        """
        Export report as Markdown using Jinja2 template.
        
        Returns:
            Rendered Markdown string
        """
        template = self.env.get_template("report_markdown.jinja2")
        md = template.render(report=bundle)
        logger.info(f"Markdown export: {len(md)} chars")
        return md

    # ─── 3. PDF Export (Puppeteer/playwright) with fallbacks ────────────

    async def export_pdf(self, bundle: ReportBundle) -> str:
        """
        Export report as PDF using cascading fallback strategy:
          1. Playwright (Puppeteer-compatible) — Browser-based PDF generation
          2. WeasyPrint — Pure Python HTML→PDF (no browser needed)
          3. Markdown-to-PDF — Last resort via mdtohtml→weasyprint
        
        Returns:
            Storage key of the generated PDF
        
        Raises:
            Exception: If all PDF engines fail
        """
        pdf_key = f"reports/{bundle.report_id}/report.pdf"
        errors = []
        
        # ── Engine 1: Playwright (Puppeteer-compatible) ──────────────────
        try:
            from playwright.async_api import async_playwright
            
            html_content = await self.export_html(bundle)
            
            async with async_playwright() as p:
                browser = await p.chromium.launch()
                page = await browser.new_page()
                await page.set_content(html_content, wait_until="networkidle")
                pdf_bytes = await page.pdf(
                    format="A4",
                    print_background=True,
                    margin={"top": "0.5in", "bottom": "0.5in", "left": "0.75in", "right": "0.75in"},
                )
                await browser.close()
            
            await self._store_export(pdf_key, pdf_bytes, "application/pdf")
            logger.info(f"PDF export (playwright): {len(pdf_bytes)} bytes")
            return pdf_key
            
        except ImportError:
            errors.append("playwright not installed")
        except Exception as e:
            errors.append(f"playwright failed: {e}")
        
        # ── Engine 2: WeasyPrint (pure Python fallback) ──────────────────
        try:
            import weasyprint
            
            html_content = await self.export_html(bundle)
            pdf_bytes = weasyprint.HTML(string=html_content).write_pdf()
            
            await self._store_export(pdf_key, pdf_bytes, "application/pdf")
            logger.info(f"PDF export (weasyprint): {len(pdf_bytes)} bytes")
            return pdf_key
            
        except ImportError:
            errors.append("weasyprint not installed")
        except Exception as e:
            errors.append(f"weasyprint failed: {e}")
        
        # ── Engine 3: fpdf2 text-based PDF (last resort, pure Python) ────
        try:
            from fpdf import FPDF
            
            md_content = await self.export_markdown(bundle)
            
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.set_font("Helvetica", "B", 16)
            
            # Title
            for line in md_content.split("\n"):
                line = line.strip()
                if line.startswith("# "):
                    pdf.set_font("Helvetica", "B", 16)
                    pdf.cell(0, 10, line.replace("# ", ""), new_x="LMARGIN", new_y="NEXT")
                    pdf.set_font("Helvetica", "B", 12)
                elif line.startswith("## "):
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.cell(0, 8, line.replace("## ", ""), new_x="LMARGIN", new_y="NEXT")
                elif line.startswith("| "):
                    # Simple table handling
                    pdf.set_font("Helvetica", "", 9)
                    cells = [c.strip() for c in line.split("|") if c.strip()]
                    for cell in cells:
                        pdf.cell(40, 6, cell[:30], border=1)
                    pdf.ln()
                elif line and not line.startswith("---"):
                    pdf.set_font("Helvetica", "", 10)
                    pdf.multi_cell(0, 5, line[:200])
            
            pdf_bytes = pdf.output()
            await self._store_export(pdf_key, pdf_bytes, "application/pdf")
            logger.info(f"PDF export (fpdf2): {len(pdf_bytes)} bytes")
            return pdf_key
            
        except ImportError:
            errors.append("fpdf2 not installed")
        except Exception as e:
            errors.append(f"fpdf2 failed: {e}")
        
        # All engines failed
        error_msg = "; ".join(errors)
        logger.error(f"PDF export failed — all 3 engines exhausted: {error_msg}")
        raise RuntimeError(f"PDF export unavailable: {error_msg}")

    # ─── 4. Excel Export (OpenPyXL) ───────────────────────────────────────

    async def export_excel(self, bundle: ReportBundle) -> str:
        """
        Export report data as Excel workbook using OpenPyXL.
        
        Creates a workbook with:
          - Sheet 1: "Summary" — Report overview and confidence scores
          - Sheet 2-N: Individual section sheets with content
          - Last Sheet: "Audit Trail" — Full transformation history
        
        Returns:
            Storage key of the generated Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # ── Sheet 1: Summary ──────────────────────────────────────────
            ws = wb.active
            ws.title = "Summary"
            ws.merge_cells("A1:D1")
            ws["A1"] = bundle.title
            ws["A1"].font = Font(bold=True, size=16)
            
            ws["A3"] = "Report ID"
            ws["B3"] = bundle.report_id
            ws["A4"] = "Generated At"
            ws["B4"] = bundle.generated_at.isoformat()
            ws["A5"] = "Overall Confidence"
            ws["B5"] = f"{bundle.overall_confidence:.0%}"
            
            # Section overview table
            ws["A7"] = "Section"
            ws["B7"] = "Confidence"
            ws["C7"] = "Content Length"
            ws["D7"] = "Badge"
            for cell in [ws["A7"], ws["B7"], ws["C7"], ws["D7"]]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            for i, section in enumerate(bundle.sections, start=8):
                ws[f"A{i}"] = section.title
                ws[f"B{i}"] = f"{section.confidence:.0%}"
                ws[f"C{i}"] = len(section.content)
                
                if section.confidence >= 0.90:
                    badge = "🟢 Auto-Approved"
                elif section.confidence >= 0.70:
                    badge = "🟡 Manual Approval"
                elif section.confidence >= 0.50:
                    badge = "🟠 Review Required"
                else:
                    badge = "🔴 Advisory Only"
                ws[f"D{i}"] = badge
            
            # ── Sheet 2+: Individual Sections ─────────────────────────────
            for section in bundle.sections:
                ws = wb.create_sheet(title=section.title[:31])  # Excel sheet name max 31 chars
                ws.merge_cells("A1:D1")
                ws["A1"] = section.title
                ws["A1"].font = Font(bold=True, size=14)
                
                ws["A3"] = "Confidence"
                ws["B3"] = f"{section.confidence:.0%}"
                
                ws["A5"] = "Content"
                ws["A6"] = section.content
                ws["A6"].alignment = Alignment(wrap_text=True)
                ws.column_dimensions["A"].width = 100
            
            # ── Last Sheet: Audit Trail ───────────────────────────────────
            ws = wb.create_sheet(title="Audit Trail")
            ws["A1"] = "Transformation Audit Trail"
            ws["A1"].font = Font(bold=True, size=14)
            
            headers = ["Step", "Column", "Description", "Timestamp", "Status"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            for row_idx, entry in enumerate(bundle.audit_trail, start=4):
                ws.cell(row=row_idx, column=1, value=entry.get("step", ""))
                ws.cell(row=row_idx, column=2, value=entry.get("column", ""))
                ws.cell(row=row_idx, column=3, value=entry.get("description", ""))
                ws.cell(row=row_idx, column=4, value=entry.get("timestamp", ""))
                ws.cell(row=row_idx, column=5, value=entry.get("status", ""))
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Store
            xlsx_key = f"reports/{bundle.report_id}/report.xlsx"
            await self._store_export(
                xlsx_key, output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            logger.info(f"Excel export: {len(output.getvalue())} bytes")
            return xlsx_key
            
        except ImportError:
            logger.warning("openpyxl not installed. Excel export unavailable.")
            raise
        except Exception as e:
            logger.error(f"Excel generation failed: {e}")
            raise

    # ─── Storage Helpers ──────────────────────────────────────────────────

    async def _store_export(
        self,
        key: str,
        data: bytes,
        content_type: str,
    ) -> str:
        """
        Store an export file in S3/MinIO with local fallback.
        
        Attempts S3/MinIO first. If that fails, falls back to local
        filesystem at /tmp/autoinsight-exports/{key}.
        
        Args:
            key: Storage key path
            data: File bytes
            content_type: MIME type
        
        Returns:
            Storage key
        """
        def _write_local(path: Path, content: bytes) -> None:
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, "wb") as f:
                f.write(content)
        
        try:
            if hasattr(storage_manager, '_s3_client') and storage_manager._s3_client:
                storage_manager._s3_client.put_object(
                    Bucket=storage_manager._bucket,
                    Key=key,
                    Body=data,
                    ContentType=content_type,
                )
            else:
                _write_local(Path("/tmp/autoinsight-exports") / key, data)
        except Exception as e:
            logger.warning(f"S3 export failed for {key}: {e}. Using local fallback.")
            _write_local(Path("/tmp/autoinsight-exports") / key, data)
        
        return key

    async def _persist_report(
        self,
        bundle: ReportBundle,
        export_urls: Dict[str, str],
    ) -> None:
        """
        Persist report metadata to PostgreSQL.
        
        Args:
            bundle: Complete ReportBundle
            export_urls: Dict of export format URLs
        """
        try:
            await insert_one("reports", {
                "report_id": bundle.report_id,
                "title": bundle.title,
                "overall_confidence": bundle.overall_confidence,
                "section_count": len(bundle.sections),
                "export_urls": json.dumps(export_urls),
                "audit_trail": json.dumps(bundle.audit_trail),
                "created_at": bundle.generated_at.isoformat(),
            }, returning="report_id")
        except Exception as e:
            logger.warning(f"Failed to persist report metadata: {e}")

    # ─── Helper Methods ───────────────────────────────────────────────────

    def _compute_overall_confidence(
        self,
        sections: List[ReportSection],
    ) -> float:
        """Compute overall confidence across all sections."""
        if not sections:
            return 0.0
        return round(
            sum(s.confidence for s in sections) / len(sections),
            4,
        )

    def _build_viz_payload(
        self,
        sections: List[ReportSection],
    ) -> Dict[str, Any]:
        """Build visualization payload from section chart hints."""
        all_charts = []
        for section in sections:
            all_charts.extend(section.chart_hints)
        
        return {
            "charts": all_charts,
            "theme": "light",
            "interactivity": {
                "zoom": True,
                "pan": True,
                "hover_tooltips": True,
                "drill_down": True,
            },
        }
